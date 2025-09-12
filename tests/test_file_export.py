# tests/test_file_export.py
import json
from pathlib import Path
import pytest

from core.file_exporter import export_file
from config import ALLOWED_FILE_TYPES_OUT, SAVE_DIR

SESSION = "test_session"


@pytest.mark.parametrize("ext", ["txt", "py", "md"])
def test_export_textlike(ext, tmp_path, monkeypatch):
    monkeypatch.setattr("config.SAVE_DIR", tmp_path)
    path = export_file(SESSION, "testfile", "Hello World", ext)
    assert path.exists()
    assert path.suffix == f".{ext}"
    assert "Hello" in path.read_text(encoding="utf-8")


def test_export_json(tmp_path, monkeypatch):
    monkeypatch.setattr("config.SAVE_DIR", tmp_path)
    data = {"a": 1, "b": 2}
    path = export_file(SESSION, "data", data, "json")
    assert path.exists()
    with open(path, encoding="utf-8") as f:
        content = json.load(f)
    assert content == data


def test_export_csv(tmp_path, monkeypatch):
    monkeypatch.setattr("config.SAVE_DIR", tmp_path)
    rows = [["col1", "col2"], [1, 2], [3, 4]]
    path = export_file(SESSION, "table", rows, "csv")
    text = path.read_text(encoding="utf-8")
    assert "col1,col2" in text
    assert "3,4" in text


def test_export_docx(tmp_path, monkeypatch):
    from docx import Document
    monkeypatch.setattr("config.SAVE_DIR", tmp_path)
    path = export_file(SESSION, "doc", "Texte docx", "docx")
    doc = Document(path)
    text = "\n".join(p.text for p in doc.paragraphs)
    assert "Texte docx" in text


def test_export_pdf(tmp_path, monkeypatch):
    monkeypatch.setattr("config.SAVE_DIR", tmp_path)
    path = export_file(SESSION, "doc", "Texte PDF", "pdf")
    assert path.exists()
    # Pas de lecture PDF (lourd), on teste juste l’existence et taille
    assert path.stat().st_size > 100


def test_export_xlsx(tmp_path, monkeypatch):
    import openpyxl
    monkeypatch.setattr("config.SAVE_DIR", tmp_path)
    rows = [["a", "b"], [1, 2]]
    path = export_file(SESSION, "sheet", rows, "xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    assert ws["A1"].value == "a"
    assert ws["B2"].value == 2
